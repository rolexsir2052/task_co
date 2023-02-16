from django.shortcuts import render
from app.models import *
from django.http import HttpResponse
# Create your views here.
def insert_data(request):
    import openpyxl
    w=openpyxl.load_workbook('C:\\Users\\pavan\\OneDrive\\Documents\\Django\\rolex\\Scripts\\project1000\\app\\n_1.xlsx')
    w1=w["Sheet1"]
    for row in w1.iter_rows(min_row=2,values_only=True):
        N=Normal.objects.get_or_create(name=row[0],email=row[1],salary=row[2])[0]
        N.save()
    return HttpResponse('inserted data')

def insert_data1(request):
    import openpyxl
    workbook=openpyxl.load_workbook('C:\\Users\\pavan\\OneDrive\\Documents\\Django\\rolex\\Scripts\\project1000\\app\\g_1.xlsx')
    worksheet=workbook["Sheet1"]
    for row in worksheet.iter_rows(min_row=2,values_only=True):
        N=Google.objects.get_or_create(name=row[0],email=row[1],salary=row[2])[0]
        N.save()
    return HttpResponse('inserted ccefully')

def display_data(request):
    normal_objs = Normal.objects.all()
    google_objs = Google.objects.all()
    obj = []
    for normal_obj in normal_objs:
        if google_objs.filter(salary=normal_obj.salary).exists():
            obj.append(normal_obj)
    context = {'obj': obj}
    return render(request, 'data.html', context)

